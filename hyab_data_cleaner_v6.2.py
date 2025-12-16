"""
HYAB Data Cleaner + Intelligence Dashboard v3.0
Generates downloadable HTML dashboard with all of Victor's requested charts

Features:
- Order Book: Cleaning, FX conversion, aging alerts
- Sales: Monthly cleanup with Top 20 analysis
- Intelligence: Generates beautiful HTML dashboard with:
  - LTM trend chart (rolling 12 months over time)
  - Monthly sales bar chart
  - YoY same-month comparison
  - Revenue bridge visualization
  - Customer cohorts (Churned, Declining, Growing, New)
  - Top 20 Customers AND Articles
  - AI-generated commentary (requires Claude API key)
"""

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from datetime import datetime
from io import BytesIO
from collections import defaultdict
import json
import requests


def classify_customer_industry(name):
    """Classify customer by industry based on name keywords"""
    if not name or str(name).strip() == '':
        return 'Other/General'
    
    name_lower = str(name).lower()
    
    industry_keywords = {
        'Automotive/EV': ['volvo', 'scania', 'autoliv', 'automotive', 'car-o-liner', 'motor', 'vehicle', 'polestar'],
        'Manufacturing/Industrial': ['manufacturing', 'industri', 'produktion', 'verkstad', 'maskin', 'tool', 'sandvik', 'atlas copco', 'skf', 'abb', 'seco tool'],
        'Electronics/Tech': ['elektr', 'electronic', 'tech', 'sensor', 'circuit', 'pcb', 'kitron', 'zollner', 'gpv', 'keba', 'automation', 'amada'],
        'Medical/Healthcare': ['medical', 'medic', 'health', 'pharma', 'hospital', 'dental', 'medi', 'gambro'],
        'Energy/Wind': ['energy', 'energi', 'wind', 'solar', 'power', 'kraft', 'vattenfall', 'nibe', 'heat pump', 'linak'],
        'Mining/Steel': ['mining', 'gruv', 'steel', 'stål', 'metall', 'metal', 'ssab', 'lkab', 'boliden', 'metso'],
        'Food/Packaging': ['food', 'livsmedel', 'tetra pak', 'packaging', 'förpackning', 'metos'],
        'Construction/Building': ['bygg', 'construction', 'building', 'fastighet', 'property'],
        'Defense/Aerospace': ['defense', 'försvar', 'military', 'aerospace', 'flyg', 'saab'],
        'Logistics/Transport': ['transport', 'logist', 'shipping', 'frakt', 'truckcam'],
    }
    
    for industry, keywords in industry_keywords.items():
        for kw in keywords:
            if kw in name_lower:
                return industry
    return 'Other/General'


def analyze_industries(cust_data, curr_ltm_col, prev_ltm_col):
    """Analyze customer cohorts by industry"""
    results = []
    industry_stats = defaultdict(lambda: {'curr': 0, 'prev': 0, 'churned_rev': 0, 'growing_rev': 0, 'declining_rev': 0, 'new_rev': 0, 'count': 0})
    
    for _, row in cust_data.iterrows():
        name = row.get('Kund', '')
        curr = float(row.get(curr_ltm_col, 0) or 0)
        prev = float(row.get(prev_ltm_col, 0) or 0)
        
        if curr == 0 and prev == 0:
            continue
            
        industry = classify_customer_industry(name)
        industry_stats[industry]['curr'] += curr
        industry_stats[industry]['prev'] += prev
        industry_stats[industry]['count'] += 1
        
        if curr == 0 and prev > 0:
            industry_stats[industry]['churned_rev'] += prev
        elif curr > 0 and prev == 0:
            industry_stats[industry]['new_rev'] += curr
        elif curr > prev:
            industry_stats[industry]['growing_rev'] += (curr - prev)
        elif curr < prev:
            industry_stats[industry]['declining_rev'] += (prev - curr)
    
    # Format results
    for industry, stats in sorted(industry_stats.items(), key=lambda x: x[1]['curr'], reverse=True):
        if stats['curr'] > 0 or stats['prev'] > 0:
            change_pct = ((stats['curr'] - stats['prev']) / stats['prev'] * 100) if stats['prev'] > 0 else 0
            results.append({
                'industry': industry,
                'curr_ltm': stats['curr'],
                'prev_ltm': stats['prev'],
                'change_pct': change_pct,
                'churned_rev': stats['churned_rev'],
                'new_rev': stats['new_rev'],
                'count': stats['count']
            })
    
    return results


def format_industry_analysis(industry_data):
    """Format industry analysis for AI prompt"""
    if not industry_data:
        return "No industry data available"
    
    lines = []
    lines.append("Industry Performance (by current LTM revenue):")
    
    for ind in industry_data[:10]:  # Top 10 industries
        trend = "📈" if ind['change_pct'] > 0 else "📉" if ind['change_pct'] < 0 else "➡️"
        lines.append(f"  {trend} {ind['industry']}: {ind['curr_ltm']:,.0f} SEK ({ind['change_pct']:+.1f}% YoY, {ind['count']} customers)")
        if ind['churned_rev'] > 50000:
            lines.append(f"      ⚠️ Lost {ind['churned_rev']:,.0f} SEK to churn")
    
    # Summary of concerning trends
    declining = [i for i in industry_data if i['change_pct'] < -10 and i['curr_ltm'] > 100000]
    if declining:
        lines.append("\n⚠️ SECTORS WITH MAJOR DECLINE (>10% drop):")
        for ind in declining:
            lines.append(f"  - {ind['industry']}: {ind['change_pct']:+.1f}% ({ind['churned_rev']:,.0f} SEK churned)")
    
    growing = [i for i in industry_data if i['change_pct'] > 15 and i['curr_ltm'] > 100000]
    if growing:
        lines.append("\n✅ SECTORS WITH STRONG GROWTH (>15% increase):")
        for ind in growing:
            lines.append(f"  - {ind['industry']}: {ind['change_pct']:+.1f}%")
    
    return "\n".join(lines)


def generate_ai_commentary(api_key, chart_data):
    """Generate AI commentary for dashboard charts using Claude API"""
    if not api_key:
        return {}
    
    prompt = f"""You are a business intelligence analyst reviewing HYAB Magneter AB's sales dashboard.

**⚠️ CRITICAL: ACCURACY REQUIREMENTS**
- ONLY make statements that are directly supported by the data below
- Do NOT invent customer relationships or assume companies are customers
- Clearly distinguish between: EXISTING customers, CHURNED customers, and companies that have NEVER been customers
- Do NOT suggest "winning back" companies that were never customers - that's NEW BUSINESS development

=== SECTION 1: COMPANY PROFILE ===

HYAB Magneter AB: Swedish industrial magnet supplier, founded 1964, based in Bromma.
2000+ articles in stock. B2B focus across multiple industrial sectors.
Total LTM Revenue: {chart_data['total_ltm']:,.0f} SEK | Active Customers: {chart_data['active_customers']}

=== SECTION 2: HYAB'S ACTUAL MARKET SECTORS (from customer analysis) ===

**These are the sectors HYAB actually serves - this is their Serviceable Obtainable Market:**

1. ELECTRONICS MANUFACTURING (EMS) - Contract manufacturers
   Key accounts: Stera Technologies (502K), Zollner Elektronik (433K declining), Kitron (291K declining), NOTE group, Elpac
   
2. INDUSTRIAL AUTOMATION & ACTUATORS - Major sector
   Key accounts: Linak A/S (1.12M, growing +16%), Keba Sweden (264K, growing), OEM Automatic (963K, growing), Atlas Copco (342K, growing), AMADA (220K, declining)
   
3. ELEVATORS & LIFT EQUIPMENT
   Key accounts: Eurosign LIFTPLAQ (1.89M - TOP CUSTOMER, +70%), Motala Hissar, Cibes Lift, Kone (small ~22K)
   
4. HVAC / HEAT PUMPS
   Key accounts: NIBE AB (1.20M, growing +50% - SUCCESS), SwedeHeat, Lindab Ventilation
   
5. FOOD PROCESSING
   Key accounts: Tetra Pak (422K, growing +142% - SUCCESS), Metos (345K, declining)
   
6. PRECISION METALWORKING
   Key accounts: Laserkraft (270K), various plåt/metalworking shops
   
7. AUTOMOTIVE (small presence)
   Key accounts: Volvo entities (~73K total, DECLINING), Scania (~11K, declining), Car-O-Liner
   
8. DEFENCE/AEROSPACE
   Key accounts: SAAB entities (small)

=== SECTION 3: DASHBOARD METRICS ===

Summary:
- Current LTM: {chart_data['total_ltm']:,.0f} SEK | Prior LTM: {chart_data['prev_ltm']:,.0f} SEK
- YoY Change: {chart_data['yoy_pct']:+.1f}% ({chart_data['yoy_chg']:+,.0f} SEK)
- Churned: {chart_data['churned_count']} customers ({chart_data['churn_loss']:,.0f} SEK lost)
- New: {chart_data['new_count']} customers ({chart_data['new_gain']:,.0f} SEK gained)
- Growing: {chart_data['growing_count']} ({chart_data['growth_gain']:,.0f} SEK) | Declining: {chart_data['declining_count']} ({chart_data['decline_loss']:,.0f} SEK)

LTM Trend: {chart_data['ltm_trend_summary']}

Top Churned: {chart_data['top_churned']}

At-Risk (declining 3+ months): {chart_data['at_risk_summary']}

Top Growing: {chart_data['top_growing']}

Churn Timeline: {chart_data['churn_timeline_summary']}

Top 20 Concentration: {chart_data['concentration_pct']:.1f}% of revenue

Industry Performance: {chart_data.get('industry_analysis', 'Not available')}

=== SECTION 4: VERIFIED CUSTOMER STATUS ===

✅ GROWING ACCOUNTS (protect & expand):
- Eurosign LIFTPLAQ: 1.89M (+775K, +70%) - elevators
- NIBE AB: 1.20M (+401K, +50%) - heat pumps
- Linak A/S: 1.12M (+156K, +16%) - actuators
- OEM Automatic: 963K (+308K, +47%) - automation
- Tetra Pak: 422K (+248K, +142%) - food processing
- Atlas Copco: 342K (+82K, +32%) - industrial

⚠️ DECLINING ACCOUNTS (stop the bleed):
- Zollner Elektronik: 433K (-176K, -29%) - EMS
- Metos: 345K (-110K, -24%) - food equipment
- Kitron: 291K (-167K, -36%) - EMS
- AMADA: 220K (-183K, -45%) - automation
- Sandvik Coromant: 123K (-75K, -38%) - industrial
- Volvo entities: 73K total (declining)

❌ CHURNED (win-back targets - these WERE customers):
- Metso Sweden AB: Lost 1,032,823 SEK (last order Nov 2024) - PRIORITY
- Jan Gårdefelt: Lost 463,944 SEK
- Starclip Elektrokabel: Lost 306,640 SEK

🆕 NEVER BEEN CUSTOMERS (new business development - do NOT say "win back"):
- LKAB, Boliden, Stegra (mining/steel)
- Polestar, Northvolt (EV)
- Danfoss, Grundfos, Vestas (Danish energy)
- SKF, Wärtsilä (heavy industry)

=== SECTION 5: MARKET CONTEXT (background only - do not confuse with customer data) ===

Nordic magnet market growing ~8% CAGR. HYAB growing 2.8% = losing relative share.
Competitor KAMIC Group: SEK 190M+ (7x HYAB size after acquisitions).
Growth sectors HYAB has LIMITED presence in: EV manufacturing, wind energy, mining.
Growth sectors HYAB is STRONG in: HVAC/heat pumps, actuators, elevators.

=== OUTPUT FORMAT ===

Provide analysis in JSON. Be factually accurate - only reference verified customer relationships.
**IMPORTANT: All values must be plain text strings, NOT nested objects or arrays (except strategic_recommendations which can be a list).**

{{
    "summary": "2-3 sentence health assessment based on actual metrics and sector performance",
    "ltm_trend": "LTM trend analysis as a paragraph",
    "monthly_sales": "Monthly pattern observations as a paragraph", 
    "yoy_comparison": "YoY performance analysis as a paragraph",
    "decomposition": "What's driving revenue changes as a paragraph",
    "churn_timeline": "When/why customers leaving as a paragraph",
    "revenue_bridge": "Revenue movement analysis as a paragraph",
    "cohorts": "Customer cohort health as a paragraph",
    "industry_analysis": "Which of HYAB's ACTUAL sectors are performing well vs struggling - as a paragraph of prose, NOT a nested object",
    "at_risk": "At-risk customers with specific retention actions as a paragraph",
    "top_customers": "Portfolio concentration and health analysis as a paragraph",
    "strategic_recommendations": "5-6 prioritized actions as a list: (1) PROTECT growing accounts (NIBE, Linak, Eurosign, Tetra Pak, Atlas Copco), (2) STOP DECLINE in EMS sector (Zollner, Kitron, AMADA), (3) WIN-BACK Metso (1M SEK opportunity - they WERE a customer), (4) INVESTIGATE why automotive is weak despite having Volvo/Scania, (5) Any genuine new business opportunities based on current sector strengths"
}}"""

    try:
        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "content-type": "application/json",
                "anthropic-version": "2023-06-01"
            },
            json={
                "model": "claude-opus-4-5-20251101",
                "max_tokens": 4000,
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=90
        )
        
        if response.status_code == 200:
            result = response.json()
            content = result['content'][0]['text']
            json_match = re.search(r'\{[\s\S]*\}', content)
            if json_match:
                return json.loads(json_match.group())
        else:
            st.warning(f"API error: {response.status_code}")
            return {}
    except Exception as e:
        st.warning(f"Commentary generation failed: {str(e)}")
        return {}

st.set_page_config(page_title="HYAB Data Cleaner", page_icon="📊", layout="centered", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Libre+Baskerville:ital@1&family=Source+Sans+3:wght@400;500;600&display=swap');

:root {
    --navy: #1B2A41;
    --navy-dark: #141F30;
    --grey-warm: #F7F6F4;
    --grey-border: #E2E0DC;
    --grey-text: #6B7280;
    --teal: #4A9BA8;
    --success: #2D6A4F;
    --success-light: #D8F3DC;
}

#MainMenu, footer, header, .stDeployButton {display: none !important;}
section[data-testid="stSidebar"] {display: none;}
.stApp {background: var(--grey-warm);}

/* Center the main content */
.main .block-container {
    max-width: 900px;
    padding-top: 1rem;
    padding-left: 2rem;
    padding-right: 2rem;
}

.hyab-header {
    background: var(--navy);
    padding: 14px 24px;
    margin: -1rem -2rem 1.5rem -2rem;
    display: flex;
    align-items: center;
    gap: 12px;
    font-family: 'Source Sans 3', sans-serif;
}
.hyab-header .brand-mark {color: var(--teal); font-size: 22px;}
.hyab-header .brand-text {color: white; font-size: 13px; font-weight: 600; letter-spacing: 0.12em;}
.hyab-header .brand-divider {color: rgba(255,255,255,0.25); margin: 0 4px;}
.hyab-header .brand-page {color: rgba(255,255,255,0.7); font-size: 13px;}

.page-title {
    font-family: 'Libre Baskerville', Georgia, serif;
    font-size: 26px;
    font-style: italic;
    font-weight: 400;
    color: var(--navy);
    margin: 0 0 20px 0;
}

.section-label {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--grey-text);
    margin: 0 0 10px 0;
}

div[data-testid="stRadio"] > div {
    flex-direction: row !important;
    gap: 0 !important;
    background: white;
    border: 1px solid var(--grey-border);
    padding: 3px;
    width: fit-content;
}
div[data-testid="stRadio"] > div > label {
    padding: 8px 18px !important;
    margin: 0 !important;
    border-radius: 0 !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    background: transparent !important;
}
div[data-testid="stRadio"] > div > label[data-checked="true"] {
    background: var(--navy) !important;
    color: white !important;
}
div[data-testid="stRadio"] > label {display: none !important;}

.stButton > button {
    background: var(--navy) !important;
    color: white !important;
    border: none !important;
    border-radius: 0 !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    height: 48px !important;
    width: 100% !important;
}
.stButton > button:hover {background: var(--navy-dark) !important;}
.stButton > button:disabled {background: var(--grey-border) !important; color: var(--grey-text) !important;}

.stDownloadButton > button {
    background: var(--success) !important;
    color: white !important;
    border: none !important;
    border-radius: 0 !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    height: 48px !important;
}

div[data-testid="stMetric"] {
    background: white;
    padding: 16px;
    border-left: 3px solid var(--navy);
    border-radius: 0;
}
div[data-testid="stMetric"] label {
    font-size: 11px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
    color: var(--grey-text) !important;
}
div[data-testid="stMetric"] [data-testid="stMetricValue"] {
    font-size: 24px !important;
    font-weight: 600 !important;
    color: var(--navy) !important;
}

.success-banner {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 10px 14px;
    background: var(--success-light);
    color: var(--success);
    font-size: 12px;
    font-weight: 600;
    letter-spacing: 0.05em;
    text-transform: uppercase;
    margin-bottom: 16px;
}

hr {border-color: var(--grey-border) !important;}
</style>

<div class="hyab-header">
    <span class="brand-mark">‹</span>
    <span class="brand-text">HYAB</span>
    <span class="brand-divider">/</span>
    <span class="brand-page">Data Cleaner v3.0</span>
</div>
""", unsafe_allow_html=True)

def _page_title(t): st.markdown(f'<h1 class="page-title">{t}</h1>', unsafe_allow_html=True)
def _section_label(t): st.markdown(f'<p class="section-label">{t}</p>', unsafe_allow_html=True)
def _success_banner(): st.markdown('<div class="success-banner">✓ Done</div>', unsafe_allow_html=True)

DEFAULT_FX = {'SEK': 1.0, 'EUR': 11.20, 'USD': 10.50, 'GBP': 13.30}

def find_sheet(wb, names):
    lower = {n.lower(): n for n in wb.sheetnames}
    for name in names:
        if name.lower() in lower: return wb[lower[name.lower()]]
    return wb[wb.sheetnames[0]] if len(wb.sheetnames) == 1 else None

def clean_amount(raw):
    if raw is None: return None, None
    m = re.match(r'([\d\s\xa0\.,]+)\s*(SEK|EUR|USD|GBP)?', str(raw).strip(), re.IGNORECASE)
    if not m: return None, None
    s = m.group(1).replace('\xa0', '').replace(' ', '')
    cur = m.group(2).upper() if m.group(2) else 'SEK'
    if re.search(r',\d{2}$', s): s = s.replace('.', '').replace(',', '.')
    else: s = s.replace(',', '')
    try: return float(s), cur
    except: return None, None

def clean_num(raw):
    if raw is None: return None
    s = str(raw).strip()
    if s in ['', 'n/a', 'None', '-']: return None
    s = s.replace('\xa0', '').replace(' ', '')
    if re.search(r',\d{2}$', s): s = s.replace('.', '').replace(',', '.')
    else: s = s.replace(',', '')
    try: return float(s)
    except: return raw

def fmt_sek(n):
    if abs(n) >= 1e6: return f"{n/1e6:.1f}M"
    if abs(n) >= 1e3: return f"{n/1e3:.0f}k"
    return f"{n:.0f}"

def fmt_num(n): return f"{n:,.0f}".replace(",", " ")

def ltm_sort_key(ltm_str):
    """Sort LTM keys chronologically (format: 'LTM YY-mon' with Swedish month names)"""
    swedish_months = {'jan':1,'feb':2,'mar':3,'apr':4,'maj':5,'jun':6,'jul':7,'aug':8,'sep':9,'okt':10,'nov':11,'dec':12}
    try:
        parts = ltm_str.replace('LTM ', '').split('-')
        year = int(parts[0]) + 2000
        month = swedish_months.get(parts[1].lower(), 1)
        return (year, month)
    except:
        return (9999, 99)

def parse_master(wb):
    data = {'articles': [], 'customers': [], 'monthly_totals': {}, 'ltm_trend': {}}
    
    if 'Försäljning per artikel' in wb.sheetnames:
        ws = wb['Försäljning per artikel']
        months, ltms, fys = {}, {}, {}
        for c in range(3, min(ws.max_column + 1, 100)):
            h = ws.cell(1, c).value
            if h is None: continue
            hs = str(h)
            if isinstance(h, datetime): months[h.strftime('%Y-%m')] = c
            elif hs.startswith('FY'): fys[hs] = c
            elif hs.startswith('LTM'): ltms[hs] = c
            elif hs == 'YTD': fys['YTD'] = c
        
        for r in range(2, ws.max_row + 1):
            art = ws.cell(r, 1).value
            if not art or art == 'Summa': continue
            a = {'artikelnr': str(art), 'artikelnamn': ws.cell(r, 2).value or '', 'monthly': {}, 'fy': {}, 'ltm': {}}
            for k, c in months.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): a['monthly'][k] = v
            for k, c in fys.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): a['fy'][k] = v
            for k, c in ltms.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): a['ltm'][k] = v
            if any(a['monthly'].values()) or any(a['fy'].values()): data['articles'].append(a)
    
    if 'Försäljning per kund' in wb.sheetnames:
        ws = wb['Försäljning per kund']
        months, ltms, fys, churn_c = {}, {}, {}, None
        for c in range(3, min(ws.max_column + 1, 100)):
            h = ws.cell(1, c).value
            if h is None: continue
            hs = str(h)
            if isinstance(h, datetime): months[h.strftime('%Y-%m')] = c
            elif hs.startswith('FY'): fys[hs] = c
            elif hs.startswith('LTM'): ltms[hs] = c
            elif hs == 'YTD': fys['YTD'] = c
            # Note: Bortfall column contains YoY change amount, NOT a churn flag
        
        for r in range(2, ws.max_row + 1):
            kund = ws.cell(r, 2).value
            if not kund or kund == 'Summa': continue
            cust = {'kund': str(kund), 'monthly': {}, 'fy': {}, 'ltm': {}}
            for k, c in months.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): cust['monthly'][k] = v
            for k, c in fys.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): cust['fy'][k] = v
            for k, c in ltms.items():
                v = ws.cell(r, c).value
                if v and isinstance(v, (int, float)): cust['ltm'][k] = v
            if any(cust['monthly'].values()) or any(cust['fy'].values()): data['customers'].append(cust)
    
    all_m = set()
    for a in data['articles']: all_m.update(a['monthly'].keys())
    for m in sorted(all_m): data['monthly_totals'][m] = sum(a['monthly'].get(m, 0) for a in data['articles'])
    
    all_l = set()
    for c in data['customers']: all_l.update(c['ltm'].keys())
    for l in sorted(all_l, key=ltm_sort_key): data['ltm_trend'][l] = sum(c['ltm'].get(l, 0) for c in data['customers'])
    
    return data

def analyze_cohorts(data, curr, prev):
    churned, declining, growing, new = [], [], [], []
    for c in data['customers']:
        cur_v = c['ltm'].get(curr, 0)
        pre_v = c['ltm'].get(prev, 0)
        chg = cur_v - pre_v
        # Churned = had revenue last year but ZERO this year
        if pre_v > 0 and cur_v == 0:
            # Find last month with revenue
            last_month = None
            sorted_months = sorted(c['monthly'].keys())
            for m in reversed(sorted_months):
                if c['monthly'].get(m, 0) > 0:
                    last_month = m
                    break
            churned.append({'kund': c['kund'], 'previous': pre_v, 'current': cur_v, 'change': chg, 'last_month': last_month})
        elif pre_v == 0 and cur_v > 0:
            new.append({'kund': c['kund'], 'current': cur_v})
        elif chg < 0:
            declining.append({'kund': c['kund'], 'previous': pre_v, 'current': cur_v, 'change': chg, 'pct': (chg/pre_v*100) if pre_v > 0 else 0})
        elif chg > 0:
            growing.append({'kund': c['kund'], 'previous': pre_v, 'current': cur_v, 'change': chg, 'pct': (chg/pre_v*100) if pre_v > 0 else 0})
    
    churned.sort(key=lambda x: x['previous'], reverse=True)
    declining.sort(key=lambda x: x['change'])
    growing.sort(key=lambda x: x['change'], reverse=True)
    new.sort(key=lambda x: x['current'], reverse=True)
    
    # Build churn timeline (group by last order month)
    churn_timeline = {}
    for c in churned:
        m = c.get('last_month') or 'Unknown'
        if m not in churn_timeline:
            churn_timeline[m] = {'count': 0, 'total': 0}
        churn_timeline[m]['count'] += 1
        churn_timeline[m]['total'] += c['previous']
    
    return {'churned': churned, 'declining': declining, 'growing': growing, 'new': new, 'churn_timeline': churn_timeline}

def get_top20_art(data, ltm_key):
    arts = [{'artikelnr': a['artikelnr'], 'artikelnamn': a['artikelnamn'], 'value': a['ltm'].get(ltm_key, 0)} for a in data['articles'] if a['ltm'].get(ltm_key, 0) > 0]
    return sorted(arts, key=lambda x: x['value'], reverse=True)[:20]

def get_top20_cust(data, curr, prev):
    custs = []
    for c in data['customers']:
        cur_v = c['ltm'].get(curr, 0)
        if cur_v > 0:
            custs.append({'kund': c['kund'], 'current': cur_v, 'previous': c['ltm'].get(prev, 0), 'change': cur_v - c['ltm'].get(prev, 0)})
    return sorted(custs, key=lambda x: x['current'], reverse=True)[:20]


def analyze_ltm_decomposition(data, num_periods=12):
    """Decompose LTM changes month-over-month into churn, decline, growth, and new"""
    all_ltms = sorted(data['ltm_trend'].keys(), key=ltm_sort_key)
    if len(all_ltms) < 2:
        return {'periods': [], 'decomposition': []}
    
    recent_ltms = all_ltms[-num_periods:] if len(all_ltms) >= num_periods else all_ltms
    decomposition = []
    
    for i in range(1, len(recent_ltms)):
        curr_ltm = recent_ltms[i]
        prev_ltm = recent_ltms[i-1]
        
        churn_loss = 0  # Had revenue in prev, zero in curr
        decline_loss = 0  # Lower revenue in curr vs prev
        growth_gain = 0  # Higher revenue in curr vs prev  
        new_gain = 0  # Zero in prev, has revenue in curr
        
        for c in data['customers']:
            curr_v = c['ltm'].get(curr_ltm, 0)
            prev_v = c['ltm'].get(prev_ltm, 0)
            
            if prev_v > 0 and curr_v == 0:
                churn_loss += prev_v
            elif prev_v == 0 and curr_v > 0:
                new_gain += curr_v
            elif curr_v < prev_v:
                decline_loss += (prev_v - curr_v)
            elif curr_v > prev_v:
                growth_gain += (curr_v - prev_v)
        
        decomposition.append({
            'period': curr_ltm,
            'total_change': data['ltm_trend'].get(curr_ltm, 0) - data['ltm_trend'].get(prev_ltm, 0),
            'churn': -churn_loss,
            'decline': -decline_loss,
            'growth': growth_gain,
            'new': new_gain
        })
    
    return {
        'periods': recent_ltms[1:],
        'decomposition': decomposition
    }


def analyze_ltm_trajectories(data, num_periods=6):
    """Analyze customer LTM trajectories over multiple periods to identify declining patterns"""
    # Get sorted LTM keys (chronological)
    all_ltms = sorted(data['ltm_trend'].keys(), key=ltm_sort_key)
    if len(all_ltms) < 3:
        return {'at_risk': [], 'consistent_decline': [], 'periods': []}
    
    # Get last N periods
    recent_ltms = all_ltms[-num_periods:] if len(all_ltms) >= num_periods else all_ltms
    
    at_risk = []  # Declining for 3+ consecutive periods
    consistent_decline = []  # Every period lower than previous
    
    for c in data['customers']:
        # Get values for recent periods
        values = [(ltm, c['ltm'].get(ltm, 0)) for ltm in recent_ltms]
        values_only = [v for _, v in values]
        
        # Skip if no significant revenue
        if max(values_only) < 50000:  # Skip small customers
            continue
        
        # Count consecutive declines from most recent
        consecutive_declines = 0
        for i in range(len(values_only) - 1, 0, -1):
            if values_only[i] < values_only[i-1]:
                consecutive_declines += 1
            else:
                break
        
        # Check if consistently declining (every period lower)
        all_declining = all(values_only[i] < values_only[i-1] for i in range(1, len(values_only)) if values_only[i-1] > 0)
        
        if consecutive_declines >= 3 and values_only[-1] > 0:
            # Calculate total decline
            peak_val = max(values_only[:-consecutive_declines]) if len(values_only) > consecutive_declines else values_only[0]
            curr_val = values_only[-1]
            decline_pct = ((peak_val - curr_val) / peak_val * 100) if peak_val > 0 else 0
            at_risk.append({
                'kund': c['kund'],
                'current': curr_val,
                'peak': peak_val,
                'decline_pct': decline_pct,
                'consecutive_months': consecutive_declines,
                'trajectory': values_only[-6:] if len(values_only) >= 6 else values_only
            })
        
        if all_declining and len(values_only) >= 4 and values_only[-1] > 0:
            consistent_decline.append({
                'kund': c['kund'],
                'start': values_only[0],
                'current': values_only[-1],
                'decline_pct': ((values_only[0] - values_only[-1]) / values_only[0] * 100) if values_only[0] > 0 else 0,
                'trajectory': values_only
            })
    
    # Sort by decline severity
    at_risk = sorted(at_risk, key=lambda x: x['decline_pct'], reverse=True)[:10]
    consistent_decline = sorted(consistent_decline, key=lambda x: x['start'] - x['current'], reverse=True)[:10]
    
    return {
        'at_risk': at_risk,
        'consistent_decline': consistent_decline,
        'periods': recent_ltms
    }


def generate_html(data, curr_ltm, prev_ltm, commentary=None):
    if commentary is None:
        commentary = {}
    
    total_ltm = data['ltm_trend'].get(curr_ltm, 0)
    prev_ltm_val = data['ltm_trend'].get(prev_ltm, 0)
    yoy_chg = total_ltm - prev_ltm_val
    yoy_pct = (yoy_chg / prev_ltm_val * 100) if prev_ltm_val > 0 else 0
    
    # Format period labels for display (e.g., "LTM 25-nov" -> "Nov 25")
    def format_ltm_label(ltm_str):
        month_map = {'jan': 'Jan', 'feb': 'Feb', 'mar': 'Mar', 'apr': 'Apr', 'maj': 'May', 'jun': 'Jun', 
                     'jul': 'Jul', 'aug': 'Aug', 'sep': 'Sep', 'okt': 'Oct', 'nov': 'Nov', 'dec': 'Dec'}
        try:
            parts = ltm_str.replace('LTM ', '').split('-')
            return f"{month_map.get(parts[1].lower(), parts[1])} {parts[0]}"
        except:
            return ltm_str
    
    # Helper to create commentary block
    def commentary_block(key):
        text = commentary.get(key, '')
        if text:
            # Handle if it's a list (structured output)
            if isinstance(text, list):
                parts = []
                for item in text:
                    if isinstance(item, dict):
                        # Extract meaningful values from dict
                        for k, v in item.items():
                            if isinstance(v, str) and len(v) > 10:
                                parts.append(v)
                    else:
                        parts.append(str(item))
                text = ' '.join(parts) if parts else str(text)
            elif isinstance(text, dict):
                # Extract values from nested dict structure
                parts = []
                def extract_strings(obj, depth=0):
                    if depth > 2:  # Limit recursion
                        return
                    if isinstance(obj, str):
                        if len(obj) > 20:  # Only meaningful strings
                            parts.append(obj)
                    elif isinstance(obj, dict):
                        for v in obj.values():
                            extract_strings(v, depth + 1)
                    elif isinstance(obj, list):
                        for item in obj:
                            extract_strings(item, depth + 1)
                extract_strings(text)
                text = ' '.join(parts) if parts else str(text)
            return f'<div class="ai-commentary"><span class="ai-icon">🤖</span> {text}</div>'
        return ''
    
    # Helper to format strategic recommendations with proper line breaks
    def format_strategic_recs(data):
        if not data:
            return ''
        
        import re
        
        # If it's a list of recommendation objects, format them nicely
        if isinstance(data, list):
            html_parts = []
            for i, rec in enumerate(data, 1):
                if isinstance(rec, dict):
                    # Handle structured recommendation object
                    priority = rec.get('priority', i)
                    action = rec.get('action', rec.get('title', f'Recommendation {i}'))
                    details = rec.get('details', rec.get('description', ''))
                    impact = rec.get('revenue_impact', rec.get('impact', ''))
                    timeline = rec.get('timeline', '')
                    
                    html = f'<br><br><strong style="color:#6366F1;">({priority})</strong> <strong>{action}</strong>'
                    if details:
                        html += f'<br>{details}'
                    if impact:
                        html += f'<br><span style="color:#10B981;font-weight:500;">→ Impact:</span> {impact}'
                    if timeline:
                        html += f'<br><span style="color:#F59E0B;font-weight:500;">→ Timeline:</span> {timeline}'
                    html_parts.append(html)
                else:
                    # Handle string item in list
                    html_parts.append(f'<br><br><strong style="color:#6366F1;">({i})</strong> {str(rec)}')
            
            result = ''.join(html_parts)
            return re.sub(r'^\s*<br><br>', '', result)  # Remove leading breaks
        
        # If it's a dict, convert to string
        if isinstance(data, dict):
            data = str(data)
        
        # Handle string input (original format)
        text = str(data)
        # Split on numbered patterns like (1), (2), 1), 2., etc.
        formatted = re.sub(r'\s*\((\d+)\)\s*', r'<br><br><strong style="color:#6366F1;">(\1)</strong> ', text)
        formatted = re.sub(r'^\s*<br><br>', '', formatted)  # Remove leading breaks
        # Also handle "Action:" and "Target:" keywords
        formatted = re.sub(r'\.\s*Action:', r'.<br><span style="color:#10B981;font-weight:500;">→ Action:</span>', formatted)
        formatted = re.sub(r'\.\s*Target:', r'.<br><span style="color:#F59E0B;font-weight:500;">→ Target:</span>', formatted)
        return formatted
    
    curr_label = format_ltm_label(curr_ltm)  # e.g., "Nov 25"
    prev_label = format_ltm_label(prev_ltm)  # e.g., "Nov 24"
    
    cohorts = analyze_cohorts(data, curr_ltm, prev_ltm)
    trajectories = analyze_ltm_trajectories(data, num_periods=6)
    ltm_decomp = analyze_ltm_decomposition(data, num_periods=12)
    churn_loss = sum(c['previous'] for c in cohorts['churned'])
    decline_loss = abs(sum(c['change'] for c in cohorts['declining']))
    growth_gain = sum(c['change'] for c in cohorts['growing'])
    new_gain = sum(c['current'] for c in cohorts['new'])
    active = len([c for c in data['customers'] if c['ltm'].get(curr_ltm, 0) > 0])
    
    top20_cust = get_top20_cust(data, curr_ltm, prev_ltm)
    top20_art = get_top20_art(data, curr_ltm)
    conc_pct = (sum(c['current'] for c in top20_cust) / total_ltm * 100) if total_ltm > 0 else 0
    
    months = sorted(data['monthly_totals'].keys())[-24:]
    m_labels = [datetime.strptime(m, '%Y-%m').strftime('%b %y') for m in months]
    m_values = [data['monthly_totals'].get(m, 0) for m in months]
    
    ltm_keys = sorted(data['ltm_trend'].keys(), key=ltm_sort_key)[-24:]
    l_labels, l_values = [], []
    month_map = {'jan': 'Jan', 'feb': 'Feb', 'mar': 'Mar', 'apr': 'Apr', 'maj': 'May', 'jun': 'Jun', 
                 'jul': 'Jul', 'aug': 'Aug', 'sep': 'Sep', 'okt': 'Oct', 'nov': 'Nov', 'dec': 'Dec'}
    for lk in ltm_keys:
        try:
            parts = lk.replace('LTM ', '').split('-')
            l_labels.append(f"{month_map.get(parts[1].lower(), parts[1])} {parts[0]}")
            l_values.append(data['ltm_trend'][lk])
        except: pass
    
    yoy_data = defaultdict(dict)
    for m in data['monthly_totals'].keys():
        try:
            dt = datetime.strptime(m, '%Y-%m')
            yoy_data[dt.month][dt.year] = data['monthly_totals'][m]
        except: pass
    years = sorted(set(y for md in yoy_data.values() for y in md.keys()))[-3:]
    yoy_ds = []
    colors = ['#64748B', '#4A9BA8', '#6366F1']
    for i, yr in enumerate(years):
        yoy_ds.append({'label': str(yr), 'data': [yoy_data[m].get(yr, 0) for m in range(1, 13)], 'backgroundColor': colors[i] if i < len(colors) else '#6366F1'})
    
    # Prepare LTM decomposition data for stacked bar chart
    decomp_labels = []
    decomp_churn = []
    decomp_decline = []
    decomp_growth = []
    decomp_new = []
    for d in ltm_decomp['decomposition'][-12:]:
        try:
            parts = d['period'].replace('LTM ', '').split('-')
            decomp_labels.append(f"{month_map.get(parts[1].lower(), parts[1])} {parts[0]}")
        except:
            decomp_labels.append(d['period'])
        decomp_churn.append(d['churn'])
        decomp_decline.append(d['decline'])
        decomp_growth.append(d['growth'])
        decomp_new.append(d['new'])
    
    # Prepare churn timeline data (when did churned customers last order)
    churn_tl = cohorts.get('churn_timeline', {})
    churn_tl_labels = []
    churn_tl_values = []
    churn_tl_counts = []
    for m in sorted(churn_tl.keys()):
        if m != 'Unknown':
            try:
                dt = datetime.strptime(m, '%Y-%m')
                churn_tl_labels.append(dt.strftime('%b %y'))
            except:
                churn_tl_labels.append(m)
            churn_tl_values.append(churn_tl[m]['total'])
            churn_tl_counts.append(churn_tl[m]['count'])
    max_val = max(total_ltm, prev_ltm_val)
    scale = 160 / max_val if max_val > 0 else 1
    
    def cohort_rows(items, cols, limit=5):
        rows = []
        for c in items[:limit]:
            if cols == 2:
                rows.append(f'<tr><td class="customer-name">{c["kund"]}</td><td class="{"positive" if c.get("current",0) > 0 else "negative"}">{fmt_num(c.get("current", c.get("previous", 0)))}</td></tr>')
            else:
                cls = "positive" if c['change'] >= 0 else "negative"
                rows.append(f'<tr><td class="customer-name">{c["kund"]}</td><td>{fmt_num(c["current"])}</td><td>{fmt_num(c["previous"])}</td><td class="{cls}">{fmt_num(c["change"])}</td></tr>')
        return '\n'.join(rows)
    
    def top_cust_rows(items):
        rows = []
        for i, c in enumerate(items[:20]):
            cls = "positive" if c['change'] >= 0 else "negative"
            rows.append(f'<tr><td class="rank-cell">{i+1}</td><td>{c["kund"]}</td><td>{fmt_num(c["current"])}</td><td class="{cls}">{("+" if c["change"]>=0 else "")}{fmt_num(c["change"])}</td><td>{c["current"]/total_ltm*100:.1f}%</td></tr>')
        return '\n'.join(rows)
    
    def trajectory_rows(items):
        rows = []
        for c in items[:7]:
            # Create mini sparkline representation
            traj = c.get('trajectory', [])
            if traj:
                max_t = max(traj) if max(traj) > 0 else 1
                bars = ''.join(['▁▂▃▄▅▆▇█'[min(7, int(v/max_t*7))] if v > 0 else '▁' for v in traj])
            else:
                bars = ''
            rows.append(f'<tr><td class="customer-name">{c["kund"]}</td><td>{fmt_num(c.get("peak", c.get("start", 0)))}</td><td>{fmt_num(c["current"])}</td><td class="negative">-{c["decline_pct"]:.0f}%</td><td style="font-family:monospace;letter-spacing:2px;">{bars}</td></tr>')
        return '\n'.join(rows)
    
    def top_art_rows(items):
        rows = []
        for i, a in enumerate(items[:20]):
            rows.append(f'<tr><td class="rank-cell">{i+1}</td><td>{a["artikelnr"]}</td><td>{(a["artikelnamn"] or "")[:40]}</td><td>{fmt_num(a["value"])}</td><td>{a["value"]/total_ltm*100:.1f}%</td></tr>')
        return '\n'.join(rows)
    
    # Pre-build churned items for the f-string (can't create dicts inside f-string)
    churned_items = [{'kund': c['kund'], 'current': -c['previous']} for c in cohorts['churned']]
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>HYAB Sales Intelligence Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
:root {{--bg-primary:#0F172A;--bg-card:#1E293B;--bg-card-hover:#334155;--text-primary:#F8FAFC;--text-secondary:#94A3B8;--text-muted:#64748B;--border:#334155;--red-churned:#DC2626;--red-churned-bg:rgba(220,38,38,0.15);--orange-declining:#F97316;--orange-declining-bg:rgba(249,115,22,0.15);--green-growing:#16A34A;--green-growing-bg:rgba(22,163,74,0.15);--teal-new:#0EA5E9;--teal-new-bg:rgba(14,165,233,0.15);--indigo-top:#6366F1;--indigo-top-bg:rgba(99,102,241,0.15);--positive:#22C55E;--negative:#EF4444;}}
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:'Inter',-apple-system,BlinkMacSystemFont,sans-serif;background:var(--bg-primary);color:var(--text-primary);line-height:1.5;}}
.dashboard{{max-width:1400px;margin:0 auto;padding:2rem;}}
.header{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:2rem;padding-bottom:1.5rem;border-bottom:1px solid var(--border);}}
.header-left h1{{font-size:1.75rem;font-weight:700;letter-spacing:-0.025em;margin-bottom:0.25rem;}}
.header-left .period{{color:var(--text-secondary);font-size:0.9rem;}}
.header-right{{text-align:right;color:var(--text-muted);font-size:0.8rem;}}
.summary-strip{{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;margin-bottom:2rem;}}
.summary-card{{background:var(--bg-card);border-radius:12px;padding:1.25rem;border:1px solid var(--border);}}
.summary-card .label{{font-size:0.75rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);margin-bottom:0.5rem;}}
.summary-card .value{{font-size:1.5rem;font-weight:700;font-family:'JetBrains Mono',monospace;}}
.summary-card .subvalue{{font-size:0.85rem;color:var(--text-secondary);margin-top:0.25rem;}}
.summary-card.highlight{{background:linear-gradient(135deg,var(--bg-card) 0%,rgba(99,102,241,0.1) 100%);border-color:var(--indigo-top);}}
.positive{{color:var(--positive);}} .negative{{color:var(--negative);}}
.chart-grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:1.5rem;margin-bottom:2rem;}}
.chart-container{{background:var(--bg-card);border-radius:12px;padding:1.5rem;border:1px solid var(--border);}}
.chart-container h3{{font-size:0.85rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);margin-bottom:1rem;}}
.chart-container .chart-wrapper{{position:relative;height:250px;width:100%;}}
.chart-section{{background:var(--bg-card);border-radius:16px;padding:1.5rem;margin-bottom:2rem;border:1px solid var(--border);}}
.chart-section h2{{font-size:1rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);margin-bottom:1rem;}}
.chart-section .chart-wrapper{{position:relative;height:200px;width:100%;}}
.revenue-bridge{{background:var(--bg-card);border-radius:16px;padding:2rem;margin-bottom:2rem;border:1px solid var(--border);}}
.revenue-bridge h2{{font-size:1rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);margin-bottom:1.5rem;}}
.bridge-container{{display:flex;align-items:flex-end;justify-content:space-between;height:200px;padding:0 1rem;}}
.bridge-item{{display:flex;flex-direction:column;align-items:center;flex:1;max-width:140px;}}
.bridge-bar{{width:60px;border-radius:4px 4px 0 0;}}
.bridge-bar.start{{background:var(--indigo-top);}} .bridge-bar.churned{{background:var(--red-churned);}}
.bridge-bar.declining{{background:var(--orange-declining);}} .bridge-bar.growing{{background:var(--green-growing);}}
.bridge-bar.new{{background:var(--teal-new);}} .bridge-bar.end{{background:linear-gradient(180deg,var(--indigo-top) 0%,#818CF8 100%);}}
.bridge-label{{margin-top:0.75rem;font-size:0.7rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-secondary);text-align:center;}}
.bridge-value{{margin-top:0.25rem;font-size:0.8rem;font-weight:600;font-family:'JetBrains Mono',monospace;}}
.bridge-connector{{flex:0.3;display:flex;align-items:center;justify-content:center;color:var(--text-muted);font-size:1.25rem;}}
.cohort-grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:1.5rem;margin-bottom:2rem;}}
.cohort-card{{background:var(--bg-card);border-radius:12px;border:1px solid var(--border);overflow:hidden;}}
.cohort-header{{padding:1.25rem;display:flex;justify-content:space-between;align-items:center;}}
.cohort-header.churned{{background:var(--red-churned-bg);border-bottom:2px solid var(--red-churned);}}
.cohort-header.declining{{background:var(--orange-declining-bg);border-bottom:2px solid var(--orange-declining);}}
.cohort-header.growing{{background:var(--green-growing-bg);border-bottom:2px solid var(--green-growing);}}
.cohort-header.new{{background:var(--teal-new-bg);border-bottom:2px solid var(--teal-new);}}
.cohort-title{{font-size:0.75rem;text-transform:uppercase;letter-spacing:0.05em;font-weight:600;}}
.cohort-header.churned .cohort-title{{color:var(--red-churned);}}
.cohort-header.declining .cohort-title{{color:var(--orange-declining);}}
.cohort-header.growing .cohort-title{{color:var(--green-growing);}}
.cohort-header.new .cohort-title{{color:var(--teal-new);}}
.cohort-count{{font-size:0.7rem;color:var(--text-secondary);margin-top:0.25rem;}}
.cohort-total{{font-family:'JetBrains Mono',monospace;font-size:1rem;font-weight:600;}}
.cohort-table{{width:100%;border-collapse:collapse;}}
.cohort-table th{{text-align:left;padding:0.75rem 1rem;font-size:0.65rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-muted);border-bottom:1px solid var(--border);font-weight:500;}}
.cohort-table th:last-child{{text-align:right;}}
.cohort-table td{{padding:0.6rem 1rem;font-size:0.8rem;border-bottom:1px solid var(--border);}}
.cohort-table td:last-child{{text-align:right;font-family:'JetBrains Mono',monospace;font-size:0.75rem;}}
.cohort-table tr:hover{{background:var(--bg-card-hover);}}
.customer-name{{color:var(--text-primary);font-weight:500;}}
.top-section{{background:var(--bg-card);border-radius:16px;border:1px solid var(--border);overflow:hidden;margin-bottom:2rem;}}
.top-header{{padding:1.25rem;background:var(--indigo-top-bg);border-bottom:2px solid var(--indigo-top);display:flex;justify-content:space-between;align-items:center;}}
.top-header h3{{font-size:0.75rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--indigo-top);font-weight:600;}}
.concentration-badge{{background:var(--indigo-top);color:white;padding:0.4rem 1rem;border-radius:20px;font-size:0.8rem;font-weight:600;}}
.top-table{{width:100%;border-collapse:collapse;}}
.top-table th{{text-align:left;padding:0.75rem 1rem;font-size:0.65rem;text-transform:uppercase;letter-spacing:0.05em;color:var(--text-muted);border-bottom:1px solid var(--border);font-weight:500;}}
.top-table td{{padding:0.6rem 1rem;font-size:0.8rem;border-bottom:1px solid var(--border);}}
.top-table tr:hover{{background:var(--bg-card-hover);}}
.rank-cell{{font-weight:600;color:var(--text-muted);width:40px;}}
.footer{{text-align:center;padding-top:1.5rem;border-top:1px solid var(--border);color:var(--text-muted);font-size:0.75rem;}}
.ai-commentary{{background:linear-gradient(135deg,rgba(99,102,241,0.1),rgba(14,165,233,0.1));border-left:3px solid #6366F1;padding:12px 16px;margin:12px 0;border-radius:0 8px 8px 0;font-size:0.85rem;color:var(--text-secondary);line-height:1.5;}}
.ai-commentary .ai-icon{{margin-right:8px;}}
.summary-commentary{{background:linear-gradient(135deg,rgba(99,102,241,0.15),rgba(14,165,233,0.1));border:1px solid rgba(99,102,241,0.3);padding:16px 20px;margin:16px 0;border-radius:12px;font-size:0.9rem;color:var(--text-primary);line-height:1.6;}}
.summary-commentary .ai-icon{{font-size:1.1rem;margin-right:10px;}}
@media(max-width:1024px){{.summary-strip,.chart-grid,.cohort-grid{{grid-template-columns:1fr;}}}}
</style>
</head>
<body>
<div class="dashboard">
<header class="header"><div class="header-left"><h1>HYAB Sales Intelligence</h1><div class="period">LTM {curr_label} vs LTM {prev_label}</div></div><div class="header-right">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}<br>Data source: Sales Master File</div></header>

<div class="summary-strip">
<div class="summary-card highlight"><div class="label">Total LTM Revenue</div><div class="value">{fmt_sek(total_ltm)} SEK</div><div class="subvalue">{fmt_num(total_ltm)} SEK</div></div>
<div class="summary-card"><div class="label">Year-over-Year</div><div class="value {'positive' if yoy_pct >= 0 else 'negative'}">{yoy_pct:+.1f}%</div><div class="subvalue">{'+' if yoy_chg >= 0 else ''}{fmt_sek(yoy_chg)} SEK</div></div>
<div class="summary-card"><div class="label">Active Customers</div><div class="value">{active}</div><div class="subvalue">{len(cohorts['churned'])} churned · {len(cohorts['new'])} new</div></div>
<div class="summary-card"><div class="label">Movement</div><div class="value">{len(cohorts['growing'])} ↑ / {len(cohorts['declining'])} ↓</div><div class="subvalue">Growing vs declining</div></div>
</div>

{f'<div class="summary-commentary"><span class="ai-icon">🤖</span> {commentary.get("summary", "")}</div>' if commentary.get("summary") else ''}

<div class="chart-grid">
<div class="chart-container"><h3>📈 Rolling LTM Trend</h3><div class="chart-wrapper"><canvas id="ltmChart"></canvas></div>{commentary_block('ltm_trend')}</div>
<div class="chart-container"><h3>📊 Monthly Sales</h3><div class="chart-wrapper"><canvas id="monthlyChart"></canvas></div>{commentary_block('monthly_sales')}</div>
</div>

<div class="chart-section"><h2>📅 Year-over-Year Comparison by Month</h2><div class="chart-wrapper"><canvas id="yoyChart"></canvas></div>{commentary_block('yoy_comparison')}</div>

<div class="chart-section"><h2>🔄 LTM Change Decomposition</h2><p style="color:#94A3B8;font-size:0.85rem;margin:-8px 0 16px 0;">Month-over-month LTM change broken down by: Churn (red), Decline (orange), Growth (green), New (teal)</p><div class="chart-wrapper"><canvas id="decompChart"></canvas></div>{commentary_block('decomposition')}</div>

<div class="chart-section"><h2>⚠️ Churn Timeline: When Customers Left</h2><p style="color:#94A3B8;font-size:0.85rem;margin:-8px 0 16px 0;">Month of LAST ORDER for churned customers (249 accounts, 3.4M SEK). Nov 24 spike = Metso's 1.03M SEK final order before leaving.</p><div class="chart-wrapper"><canvas id="churnTimelineChart"></canvas></div>{commentary_block('churn_timeline')}</div>

<div class="revenue-bridge"><h2>Revenue Bridge</h2>
<div class="bridge-container">
<div class="bridge-item"><div class="bridge-bar start" style="height:{prev_ltm_val*scale:.0f}px;"></div><div class="bridge-label">Prior Year</div><div class="bridge-value">{fmt_sek(prev_ltm_val)}</div></div>
<div class="bridge-connector">→</div>
<div class="bridge-item"><div class="bridge-bar churned" style="height:{max(churn_loss*scale*0.1,10):.0f}px;"></div><div class="bridge-label">Churn Loss</div><div class="bridge-value negative">-{fmt_sek(churn_loss)}</div></div>
<div class="bridge-item"><div class="bridge-bar declining" style="height:{max(decline_loss*scale*0.1,10):.0f}px;"></div><div class="bridge-label">Decline Loss</div><div class="bridge-value negative">-{fmt_sek(decline_loss)}</div></div>
<div class="bridge-connector">+</div>
<div class="bridge-item"><div class="bridge-bar growing" style="height:{max(growth_gain*scale*0.1,10):.0f}px;"></div><div class="bridge-label">Growth Gain</div><div class="bridge-value positive">+{fmt_sek(growth_gain)}</div></div>
<div class="bridge-item"><div class="bridge-bar new" style="height:{max(new_gain*scale*0.1,10):.0f}px;"></div><div class="bridge-label">New Customers</div><div class="bridge-value positive">+{fmt_sek(new_gain)}</div></div>
<div class="bridge-connector">=</div>
<div class="bridge-item"><div class="bridge-bar end" style="height:{total_ltm*scale:.0f}px;"></div><div class="bridge-label">Current Year</div><div class="bridge-value">{fmt_sek(total_ltm)}</div></div>
</div>
{commentary_block('revenue_bridge')}
</div>

<div class="cohort-grid">
<div class="cohort-card"><div class="cohort-header churned"><div><div class="cohort-title">⚠️ Churned Customers</div><div class="cohort-count">{len(cohorts['churned'])} accounts lost vs {prev_label}</div></div><div class="cohort-total negative">-{fmt_sek(churn_loss)} SEK</div></div><table class="cohort-table"><thead><tr><th>Customer</th><th>Lost ({prev_label})</th></tr></thead><tbody>{cohort_rows(churned_items, 2)}</tbody></table></div>
<div class="cohort-card"><div class="cohort-header declining"><div><div class="cohort-title">📉 Declining Customers</div><div class="cohort-count">{len(cohorts['declining'])} accounts declining</div></div><div class="cohort-total negative">-{fmt_sek(decline_loss)} SEK</div></div><table class="cohort-table"><thead><tr><th>Customer</th><th>{curr_label}</th><th>{prev_label}</th><th>Change</th></tr></thead><tbody>{cohort_rows(cohorts['declining'],4)}</tbody></table></div>
<div class="cohort-card"><div class="cohort-header growing"><div><div class="cohort-title">🎉 Growing Customers</div><div class="cohort-count">{len(cohorts['growing'])} accounts expanding</div></div><div class="cohort-total positive">+{fmt_sek(growth_gain)} SEK</div></div><table class="cohort-table"><thead><tr><th>Customer</th><th>{curr_label}</th><th>{prev_label}</th><th>Growth</th></tr></thead><tbody>{cohort_rows(cohorts['growing'],4,7)}</tbody></table></div>
<div class="cohort-card"><div class="cohort-header new"><div><div class="cohort-title">✨ New Customers</div><div class="cohort-count">{len(cohorts['new'])} new since {prev_label}</div></div><div class="cohort-total positive">+{fmt_sek(new_gain)} SEK</div></div><table class="cohort-table"><thead><tr><th>Customer</th><th>Revenue ({curr_label})</th></tr></thead><tbody>{cohort_rows(cohorts['new'],2,7)}</tbody></table></div>
</div>
{commentary_block('cohorts')}

{f'''<div class="top-section" style="margin-top:2rem;">
<div class="top-header" style="background:linear-gradient(135deg,rgba(139,92,246,0.15),rgba(99,102,241,0.1));border-color:#8B5CF6;">
<h3 style="color:#8B5CF6;">🏭 Industry Analysis</h3>
<span class="concentration-badge" style="background:#8B5CF6;">Sector Performance</span>
</div>
<p style="color:#6B7280;font-size:12px;padding:0 1rem;margin:8px 0;">Revenue by industry sector (classified from customer names). Highlights alignment with market growth drivers.</p>
<table class="top-table">
<thead><tr><th>Industry</th><th>Current LTM</th><th>YoY Change</th><th>Churned</th><th>Customers</th><th>Market Alignment</th></tr></thead>
<tbody>''' + ''.join([f'''<tr>
<td><strong>{ind["industry"]}</strong></td>
<td>{ind["curr_ltm"]:,.0f} SEK</td>
<td style="color:{"#10B981" if ind["change_pct"] > 0 else "#EF4444" if ind["change_pct"] < 0 else "#6B7280"};font-weight:600;">{ind["change_pct"]:+.1f}%</td>
<td style="color:#EF4444;">{f'-{ind["churned_rev"]:,.0f}' if ind["churned_rev"] > 0 else '-'}</td>
<td>{ind["count"]}</td>
<td>{"🔥 HIGH GROWTH" if ind["industry"] in ["Energy/Wind", "Automotive/EV", "Mining/Steel"] else "⚡ GROWING" if ind["industry"] in ["Defense/Aerospace", "Manufacturing/Industrial"] else "📊 STABLE"}</td>
</tr>''' for ind in data.get("industry_analysis", [])[:10]]) + '''</tbody>
</table>
{commentary_block('industry_analysis')}
</div>''' if data.get("industry_analysis") else ''}

{f'''<div class="top-section" style="margin-top:2rem;">
<div class="top-header" style="background:rgba(239,68,68,0.1);border-color:#EF4444;">
<h3 style="color:#EF4444;">🔻 At Risk: Declining for 3+ Months</h3>
<span class="concentration-badge" style="background:#EF4444;">{len(trajectories['at_risk'])} customers</span>
</div>
<p style="color:#6B7280;font-size:12px;padding:0 1rem;margin-top:-8px;">Customers with declining LTM revenue for 3 or more consecutive months</p>
<table class="top-table">
<thead><tr><th>Customer</th><th>Peak LTM</th><th>Current</th><th>Decline</th><th>Trend (6mo)</th></tr></thead>
<tbody>{trajectory_rows(trajectories['at_risk'])}</tbody>
</table>
{commentary_block('at_risk')}
</div>''' if trajectories['at_risk'] else ''}

<div class="top-section"><div class="top-header"><h3>👑 Top 20 Customers</h3><span class="concentration-badge">{conc_pct:.1f}% of Revenue</span></div><table class="top-table"><thead><tr><th>#</th><th>Customer</th><th>{curr_label}</th><th>vs {prev_label}</th><th>% of Total</th></tr></thead><tbody>{top_cust_rows(top20_cust)}</tbody></table>{commentary_block('top_customers')}</div>

<div class="top-section"><div class="top-header" style="background:rgba(74,155,168,0.15);border-color:#4A9BA8;"><h3 style="color:#4A9BA8;">📦 Top 20 Articles</h3><span class="concentration-badge" style="background:#4A9BA8;">{sum(a['value'] for a in top20_art)/total_ltm*100:.1f}% of Revenue</span></div><table class="top-table"><thead><tr><th>#</th><th>Article No</th><th>Description</th><th>{curr_label}</th><th>% of Total</th></tr></thead><tbody>{top_art_rows(top20_art)}</tbody></table></div>

{f'''<div class="top-section" style="margin-top:2rem;">
<div class="top-header" style="background:linear-gradient(135deg,rgba(99,102,241,0.15),rgba(14,165,233,0.1));border-color:#6366F1;">
<h3 style="color:#6366F1;">🚀 Strategic Recommendations</h3>
<span class="concentration-badge" style="background:#6366F1;">Based on Customer Data</span>
</div>
<div style="padding:1rem;">
<p style="color:#94A3B8;font-size:0.85rem;margin-bottom:0.5rem;"><strong>HYAB's Core Sectors:</strong> Industrial Automation/Actuators (Linak 1.1M), Elevators (Eurosign 1.9M), HVAC/Heat Pumps (NIBE 1.2M), EMS/Electronics (Stera, Zollner, Kitron), Food Processing (Tetra Pak).</p>
<p style="color:#94A3B8;font-size:0.85rem;margin-bottom:1rem;"><strong>Key Challenge:</strong> EMS sector declining (Zollner -29%, Kitron -36%, AMADA -45%). Metso churned (1.03M SEK lost). KAMIC 7x larger after acquisitions.</p>
<div class="ai-commentary" style="background:linear-gradient(135deg,rgba(99,102,241,0.1),rgba(14,165,233,0.1));border-left:4px solid #6366F1;padding:20px 24px;margin:0;border-radius:0 12px 12px 0;line-height:1.8;">
<span class="ai-icon" style="font-size:1.2rem;">🤖</span> {format_strategic_recs(commentary.get("strategic_recommendations", "Enable AI commentary with Claude API key to receive strategic recommendations."))}
</div>
</div>
</div>''' if commentary.get("strategic_recommendations") else ''}

<footer class="footer"><p>HYAB Sales Intelligence Dashboard · Generated by HYAB Data App v3.0 · {datetime.now().strftime('%Y-%m-%d')}</p></footer>
</div>

<script>
document.addEventListener('DOMContentLoaded', function() {{
    try {{
        Chart.defaults.color = '#94A3B8';
        Chart.defaults.borderColor = '#334155';
        
        var ltmLabels = {json.dumps(l_labels)};
        var ltmValues = {json.dumps(l_values)};
        var monthLabels = {json.dumps(m_labels)};
        var monthValues = {json.dumps(m_values)};
        var yoyDatasets = {json.dumps(yoy_ds)};
        var decompLabels = {json.dumps(decomp_labels)};
        var decompChurn = {json.dumps(decomp_churn)};
        var decompDecline = {json.dumps(decomp_decline)};
        var decompGrowth = {json.dumps(decomp_growth)};
        var decompNew = {json.dumps(decomp_new)};
        var churnTlLabels = {json.dumps(churn_tl_labels)};
        var churnTlValues = {json.dumps(churn_tl_values)};
        var churnTlCounts = {json.dumps(churn_tl_counts)};
        
        console.log('LTM Labels:', ltmLabels.length, 'Values:', ltmValues.length);
        console.log('Month Labels:', monthLabels.length, 'Values:', monthValues.length);
        console.log('YoY Datasets:', yoyDatasets.length);
        
        if (ltmLabels.length > 0 && ltmValues.length > 0) {{
            new Chart(document.getElementById('ltmChart'), {{
                type: 'line',
                data: {{
                    labels: ltmLabels,
                    datasets: [{{
                        label: 'LTM Sales',
                        data: ltmValues,
                        borderColor: '#6366F1',
                        backgroundColor: 'rgba(99,102,241,0.1)',
                        tension: 0.3,
                        fill: true,
                        pointRadius: 3
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{ legend: {{ display: false }} }},
                    scales: {{ y: {{ beginAtZero: false, ticks: {{ callback: function(v) {{ return (v/1000000).toFixed(1) + 'M'; }} }} }} }}
                }}
            }});
            console.log('LTM chart created');
        }}
        
        if (monthLabels.length > 0 && monthValues.length > 0) {{
            new Chart(document.getElementById('monthlyChart'), {{
                type: 'bar',
                data: {{
                    labels: monthLabels,
                    datasets: [{{
                        label: 'Monthly Sales',
                        data: monthValues,
                        backgroundColor: '#4A9BA8',
                        borderRadius: 4
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{ legend: {{ display: false }} }},
                    scales: {{ y: {{ beginAtZero: true, ticks: {{ callback: function(v) {{ return (v/1000000).toFixed(1) + 'M'; }} }} }} }}
                }}
            }});
            console.log('Monthly chart created');
        }}
        
        if (yoyDatasets.length > 0) {{
            new Chart(document.getElementById('yoyChart'), {{
                type: 'bar',
                data: {{
                    labels: ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'],
                    datasets: yoyDatasets
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{ legend: {{ position: 'top' }} }},
                    scales: {{ y: {{ beginAtZero: true, ticks: {{ callback: function(v) {{ return (v/1000000).toFixed(1) + 'M'; }} }} }} }}
                }}
            }});
            console.log('YoY chart created');
        }}
        
        if (decompLabels.length > 0) {{
            new Chart(document.getElementById('decompChart'), {{
                type: 'bar',
                data: {{
                    labels: decompLabels,
                    datasets: [
                        {{ label: 'Churn', data: decompChurn, backgroundColor: '#DC2626', stack: 'stack0' }},
                        {{ label: 'Decline', data: decompDecline, backgroundColor: '#F97316', stack: 'stack0' }},
                        {{ label: 'Growth', data: decompGrowth, backgroundColor: '#16A34A', stack: 'stack0' }},
                        {{ label: 'New', data: decompNew, backgroundColor: '#0D9488', stack: 'stack0' }}
                    ]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{ legend: {{ position: 'top' }} }},
                    scales: {{ 
                        x: {{ stacked: true }},
                        y: {{ 
                            stacked: true,
                            ticks: {{ callback: function(v) {{ return (v/1000000).toFixed(1) + 'M'; }} }}
                        }} 
                    }}
                }}
            }});
            console.log('Decomposition chart created');
        }}
        
        // Churn Timeline Chart
        if (churnTlLabels.length > 0 && churnTlValues.length > 0) {{
            new Chart(document.getElementById('churnTimelineChart'), {{
                type: 'bar',
                data: {{
                    labels: churnTlLabels,
                    datasets: [{{
                        label: 'Lost LTM Revenue',
                        data: churnTlValues,
                        backgroundColor: '#EF4444',
                        borderColor: '#DC2626',
                        borderWidth: 1
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{ 
                        legend: {{ display: false }},
                        tooltip: {{
                            callbacks: {{
                                label: function(ctx) {{
                                    var count = churnTlCounts[ctx.dataIndex];
                                    return 'Lost: ' + (ctx.parsed.y/1000).toFixed(0) + 'k SEK (' + count + ' customers)';
                                }}
                            }}
                        }}
                    }},
                    scales: {{ 
                        y: {{ 
                            ticks: {{ callback: function(v) {{ return (v/1000).toFixed(0) + 'k'; }} }}
                        }} 
                    }}
                }}
            }});
            console.log('Churn Timeline chart created');
        }}
        
        console.log('All charts initialized successfully');
    }} catch(e) {{
        console.error('Chart initialization error:', e);
        document.body.insertAdjacentHTML('afterbegin', '<div style="background:red;color:white;padding:10px;">Chart Error: ' + e.message + '</div>');
    }}
}});
</script>
</body></html>'''
    return html


# =============================================================================
# MAIN APP
# =============================================================================

_section_label("Mode")
mode = st.radio("Mode", ["Order Book", "Sales", "Intelligence"], label_visibility="collapsed")
st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

if mode == "Order Book":
    _page_title("Process Order Book")
    st.markdown('<p style="color:#6B7280;margin:-16px 0 20px 0;">Convert currencies to SEK and clean order data from Briox export.</p>', unsafe_allow_html=True)
    
    _section_label("Exchange rates")
    c1, c2, c3 = st.columns(3)
    with c1: eur = st.number_input("EUR", value=DEFAULT_FX['EUR'], step=0.01, format="%.2f")
    with c2: usd = st.number_input("USD", value=DEFAULT_FX['USD'], step=0.01, format="%.2f")
    with c3: gbp = st.number_input("GBP", value=DEFAULT_FX['GBP'], step=0.01, format="%.2f")
    fx = {'SEK': 1.0, 'EUR': eur, 'USD': usd, 'GBP': gbp}
    
    _section_label("Files")
    col1, col2 = st.columns(2)
    with col1:
        current_file = st.file_uploader("This week", type=['xlsx'], key="ob_current")
    with col2:
        previous_file = st.file_uploader("Last week (optional)", type=['xlsx'], key="ob_previous")
    
    if st.button("Process", type="primary", disabled=not current_file, use_container_width=True):
        if current_file:
            try:
                wb = openpyxl.load_workbook(current_file)
                ws = find_sheet(wb, ['Order book', 'Sheet1', 'Orders', 'Orderbook'])
                if ws is None: 
                    st.error("Could not find order book sheet")
                else:
                    orders = []
                    for r in range(2, ws.max_row + 1):
                        amt, cur = clean_amount(ws.cell(r, 6).value)
                        if amt is None: continue
                        od = ws.cell(r, 2).value
                        fakt_stat = ws.cell(r, 5).value or ''
                        orders.append({
                            'ordernr': ws.cell(r, 1).value, 
                            'orderdatum': od if isinstance(od, datetime) else None,
                            'kundnamn': ws.cell(r, 3).value,
                            'status': ws.cell(r, 4).value or '',
                            'fakt_stat': str(fakt_stat),
                            'partially_invoiced': str(fakt_stat).lower() in ['delfakt.', 'fakturerad', 'delfakturerad'],
                            'original_amount': amt,
                            'original_currency': cur,
                            'belopp_sek': round(amt * fx.get(cur, 1.0), 2)
                        })
                    st.session_state['ob_res'] = {'orders': orders, 'fx': fx}
                    st.rerun()
            except Exception as e: 
                st.error(str(e))
    
    if 'ob_res' in st.session_state:
        st.markdown("---")
        _success_banner()
        orders = st.session_state['ob_res']['orders']
        total_sek = sum(o['belopp_sek'] for o in orders)
        partial_orders = [o for o in orders if o['partially_invoiced']]
        partial_sek = sum(o['belopp_sek'] for o in partial_orders)
        
        c1, c2, c3, c4 = st.columns(4)
        with c1: st.metric("Orders", len(orders))
        with c2: st.metric("Total SEK", f"{total_sek:,.0f}")
        with c3: st.metric("Customers", len(set(o['kundnamn'] for o in orders)))
        with c4: st.metric("Partially Invoiced", f"{len(partial_orders)} ({partial_sek:,.0f} SEK)")
        
        # Filter to orders with actual value for summaries
        orders_with_value = [o for o in orders if o['belopp_sek'] > 0]
        
        # Build summaries (only orders with value)
        by_customer = {}
        for o in orders_with_value:
            k = o['kundnamn']
            if k not in by_customer:
                by_customer[k] = {'orders': 0, 'total': 0}
            by_customer[k]['orders'] += 1
            by_customer[k]['total'] += o['belopp_sek']
        customer_list = sorted(by_customer.items(), key=lambda x: x[1]['total'], reverse=True)
        
        by_month = {}
        for o in orders_with_value:
            if o['orderdatum']:
                m = o['orderdatum'].strftime('%Y-%m')
            else:
                m = 'Unknown'
            if m not in by_month:
                by_month[m] = {'orders': 0, 'total': 0}
            by_month[m]['orders'] += 1
            by_month[m]['total'] += o['belopp_sek']
        
        # === INTERACTIVE SECTION ===
        st.markdown("---")
        _section_label("Order Book Intelligence")
        
        # Filters
        col1, col2, col3 = st.columns(3)
        with col1:
            customers = ['All'] + [c[0] for c in customer_list]
            selected_customer = st.selectbox("Filter by Customer", customers, key="ob_filter_cust")
        with col2:
            fakt_options = ['All'] + sorted(set(o['fakt_stat'] for o in orders if o['fakt_stat']))
            selected_fakt = st.selectbox("Filter by Invoice Status", fakt_options, key="ob_filter_fakt")
        with col3:
            sort_options = ['Order Date (newest)', 'Order Date (oldest)', 'Amount (highest)', 'Amount (lowest)', 'Customer A-Z', 'Order Number']
            selected_sort = st.selectbox("Sort by", sort_options, key="ob_sort")
        
        # Apply filters
        filtered = orders.copy()
        if selected_customer != 'All':
            filtered = [o for o in filtered if o['kundnamn'] == selected_customer]
        if selected_fakt != 'All':
            filtered = [o for o in filtered if o['fakt_stat'] == selected_fakt]
        
        # Apply sort
        if selected_sort == 'Order Date (newest)':
            filtered = sorted(filtered, key=lambda x: x['orderdatum'] or datetime.min, reverse=True)
        elif selected_sort == 'Order Date (oldest)':
            filtered = sorted(filtered, key=lambda x: x['orderdatum'] or datetime.min)
        elif selected_sort == 'Amount (highest)':
            filtered = sorted(filtered, key=lambda x: x['belopp_sek'], reverse=True)
        elif selected_sort == 'Amount (lowest)':
            filtered = sorted(filtered, key=lambda x: x['belopp_sek'])
        elif selected_sort == 'Customer A-Z':
            filtered = sorted(filtered, key=lambda x: x['kundnamn'] or '')
        elif selected_sort == 'Order Number':
            filtered = sorted(filtered, key=lambda x: x['ordernr'] or 0, reverse=True)
        
        # Filtered summary
        filtered_total = sum(o['belopp_sek'] for o in filtered)
        st.markdown(f'<p style="color:#6B7280;margin:8px 0;">Showing {len(filtered)} orders · {filtered_total:,.0f} SEK</p>', unsafe_allow_html=True)
        
        # Orders table - use numeric values for proper sorting
        import pandas as pd
        table_data = []
        for o in filtered:
            table_data.append({
                'Ordernr': o['ordernr'],
                'Datum': o['orderdatum'].strftime('%Y-%m-%d') if o['orderdatum'] else '',
                'Kund': o['kundnamn'][:40] if o['kundnamn'] else '',
                'Fakt.stat': o['fakt_stat'],
                'Belopp SEK': o['belopp_sek']  # Keep as number for sorting
        })
        if table_data:
            df = pd.DataFrame(table_data)
            st.dataframe(
                df.style.format({'Belopp SEK': '{:,.0f}'}),
                use_container_width=True, 
                height=300
            )
        
        # Show if there are 0 SEK orders excluded
        zero_orders = [o for o in orders if o['belopp_sek'] == 0]
        if zero_orders:
            st.markdown(f'<p style="color:#9CA3AF;font-size:12px;margin-top:8px;">ℹ️ {len(zero_orders)} orders with 0 SEK excluded from summaries below (likely fully invoiced)</p>', unsafe_allow_html=True)
        
        # Monthly breakdown
        st.markdown("---")
        _section_label("Orders by Placement Month")
        st.markdown('<p style="color:#6B7280;margin:-8px 0 12px 0;">Value in order book by when orders were placed (not delivery date)</p>', unsafe_allow_html=True)
        
        month_data = []
        for m in sorted(by_month.keys(), reverse=True):
            data = by_month[m]
            month_data.append({
                'Month': m,
                'Orders': data['orders'],
                'Total SEK': f"{data['total']:,.0f}",
                '% of Total': f"{data['total']/total_sek*100:.1f}%" if total_sek > 0 else "0%"
            })
        if month_data:
            st.dataframe(month_data, use_container_width=True, height=200)
        
        # Top customers
        st.markdown("---")
        _section_label("Top Customers in Order Book")
        top_cust_data = []
        for cust, data in customer_list[:15]:
            top_cust_data.append({
                'Customer': cust[:50] if cust else '',
                'Orders': data['orders'],
                'Total SEK': f"{data['total']:,.0f}",
                '% of Total': f"{data['total']/total_sek*100:.1f}%" if total_sek > 0 else "0%"
            })
        if top_cust_data:
            st.dataframe(top_cust_data, use_container_width=True, height=300)
        
        # Month-end tracking export
        st.markdown("---")
        _section_label("Month-End Tracking")
        st.markdown('<p style="color:#6B7280;margin:-8px 0 12px 0;">Copy this line to paste into your order book tracking spreadsheet</p>', unsafe_allow_html=True)
        
        today = datetime.now()
        tracking_line = f"{today.strftime('%Y-%m-%d')}\t{total_sek:.0f}"
        st.code(tracking_line, language=None)
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f'<p style="font-size:13px;color:#6B7280;">Date: <strong>{today.strftime("%Y-%m-%d")}</strong></p>', unsafe_allow_html=True)
        with col2:
            st.markdown(f'<p style="font-size:13px;color:#6B7280;">Order Book Total: <strong>{total_sek:,.0f} SEK</strong></p>', unsafe_allow_html=True)
        
        # Download section
        st.markdown("---")
        
        # Create downloadable Excel
        output = BytesIO()
        wb_out = openpyxl.Workbook()
        
        # Summary
        ws_sum = wb_out.active
        ws_sum.title = "Summary"
        ws_sum['A1'] = "HYAB ORDER BOOK"
        ws_sum['A1'].font = Font(bold=True, size=14)
        ws_sum['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_sum['A4'] = "Total Orders"
        ws_sum['B4'] = len(orders)
        ws_sum['A5'] = "Total Value (SEK)"
        ws_sum['B5'] = total_sek
        ws_sum['B5'].number_format = '#,##0'
        ws_sum['A6'] = "Unique Customers"
        ws_sum['B6'] = len(by_customer)
        ws_sum['A7'] = "Partially Invoiced Orders"
        ws_sum['B7'] = len(partial_orders)
        ws_sum['A8'] = "Partially Invoiced Value"
        ws_sum['B8'] = partial_sek
        ws_sum['B8'].number_format = '#,##0'
        ws_sum.column_dimensions['A'].width = 25
        ws_sum.column_dimensions['B'].width = 15
        
        # All Orders
        ws_all = wb_out.create_sheet("All Orders")
        ws_all.append(['Ordernr', 'Orderdatum', 'Kundnamn', 'Fakt.stat', 'Original', 'Currency', 'Belopp SEK'])
        for o in sorted(orders, key=lambda x: x['orderdatum'] or datetime.min, reverse=True):
            ws_all.append([o['ordernr'], o['orderdatum'].strftime('%Y-%m-%d') if o['orderdatum'] else '', o['kundnamn'], o['fakt_stat'], o['original_amount'], o['original_currency'], o['belopp_sek']])
        
        # By Customer
        ws_cust = wb_out.create_sheet("By Customer")
        ws_cust['A1'] = "ORDER BOOK BY CUSTOMER"
        ws_cust['A1'].font = Font(bold=True)
        ws_cust.append([])
        ws_cust.append(['Customer', 'Orders', 'Total SEK', '% of Total'])
        for cust, data in customer_list:
            ws_cust.append([cust, data['orders'], data['total'], f"{data['total']/total_sek*100:.1f}%" if total_sek > 0 else "0%"])
        
        # By Month (only orders with value)
        ws_month = wb_out.create_sheet("By Month")
        ws_month['A1'] = "ORDERS BY PLACEMENT MONTH"
        ws_month['A2'] = f"Excludes {len(zero_orders)} orders with 0 SEK" if zero_orders else None
        ws_month['A1'].font = Font(bold=True)
        ws_month.append([])
        ws_month.append(['Month', 'Orders', 'Total SEK', '% of Total'])
        for m in sorted(by_month.keys()):
            data = by_month[m]
            ws_month.append([m, data['orders'], data['total'], f"{data['total']/total_sek*100:.1f}%" if total_sek > 0 else "0%"])
        
        # Partially Invoiced
        ws_partial = wb_out.create_sheet("Partially Invoiced")
        ws_partial['A1'] = f"PARTIALLY INVOICED ORDERS ({len(partial_orders)} orders, {partial_sek:,.0f} SEK)"
        ws_partial['A1'].font = Font(bold=True)
        ws_partial.append([])
        ws_partial.append(['Ordernr', 'Orderdatum', 'Kundnamn', 'Fakt.stat', 'Belopp SEK'])
        for o in sorted(partial_orders, key=lambda x: x['belopp_sek'], reverse=True):
            ws_partial.append([o['ordernr'], o['orderdatum'].strftime('%Y-%m-%d') if o['orderdatum'] else '', o['kundnamn'], o['fakt_stat'], o['belopp_sek']])
        
        wb_out.save(output)
        st.download_button(
            label="📥 Download Order Book Report",
            data=output.getvalue(),
            file_name=f"HYAB_OrderBook_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.markdown('<p style="color:#6B7280;font-size:12px;margin-top:8px;">Contains: Summary, All Orders, By Customer, By Month, Partially Invoiced</p>', unsafe_allow_html=True)

elif mode == "Sales":
    _page_title("Process Sales Data")
    st.markdown('<p style="color:#6B7280;margin:-16px 0 20px 0;">Parse monthly sales export for pasting into Sales_work_file.xlsx.</p>', unsafe_allow_html=True)
    
    _section_label("Exchange rates")
    c1, c2, c3 = st.columns(3)
    with c1: eur = st.number_input("EUR", value=DEFAULT_FX['EUR'], step=0.01, format="%.2f", key="sales_eur")
    with c2: usd = st.number_input("USD", value=DEFAULT_FX['USD'], step=0.01, format="%.2f", key="sales_usd")
    with c3: gbp = st.number_input("GBP", value=DEFAULT_FX['GBP'], step=0.01, format="%.2f", key="sales_gbp")
    fx = {'SEK': 1.0, 'EUR': eur, 'USD': usd, 'GBP': gbp}
    
    _section_label("Files")
    col1, col2 = st.columns(2)
    with col1:
        f = st.file_uploader("This month", type=['xlsx'], key="sales_current")
    with col2:
        f_prev = st.file_uploader("Last month (optional)", type=['xlsx'], key="sales_previous")
    
    if st.button("Process", type="primary", disabled=not f, use_container_width=True):
        if f:
            try:
                wb = openpyxl.load_workbook(f)
                ws_art = find_sheet(wb, ['Article', 'Articles', 'Artikel'])
                ws_cust = find_sheet(wb, ['Company', 'Companies', 'Företag', 'Kund'])
                
                arts = []
                custs = []
                
                if ws_art:
                    for r in range(2, ws_art.max_row + 1):
                        an = ws_art.cell(r, 1).value
                        if an is None: continue
                        s = clean_num(ws_art.cell(r, 3).value)
                        if s is None or s == 0: continue
                        arts.append({'artikelnr': str(an), 'artikelnamn': ws_art.cell(r, 2).value or '', 'summa': s})
                
                if ws_cust:
                    for r in range(2, ws_cust.max_row + 1):
                        kund = ws_cust.cell(r, 2).value
                        if kund is None: continue
                        # Sum is in column 4 for Company sheet (Col 3 is Kundtyp)
                        s = clean_num(ws_cust.cell(r, 4).value)
                        if s is None or s == 0: continue
                        custs.append({'kundnr': ws_cust.cell(r, 1).value, 'kund': str(kund), 'summa': s})
                
                st.session_state['sales_res'] = {
                    'articles': sorted(arts, key=lambda x: x['summa'], reverse=True),
                    'customers': sorted(custs, key=lambda x: x['summa'], reverse=True)
                }
                st.rerun()
            except Exception as e: 
                st.error(str(e))
    
    if 'sales_res' in st.session_state:
        st.markdown("---")
        _success_banner()
        res = st.session_state['sales_res']
        arts = res['articles']
        custs = res['customers']
        
        total_art = sum(a['summa'] for a in arts)
        total_cust = sum(c['summa'] for c in custs)
        top20_art_total = sum(a['summa'] for a in arts[:20])
        top20_cust_total = sum(c['summa'] for c in custs[:20])
        top20_art_pct = (top20_art_total / total_art * 100) if total_art > 0 else 0
        top20_cust_pct = (top20_cust_total / total_cust * 100) if total_cust > 0 else 0
        
        c1, c2, c3, c4 = st.columns(4)
        with c1: st.metric("Total Sales", f"{total_art:,.0f} SEK")
        with c2: st.metric("Articles", len(arts))
        with c3: st.metric("Customers", len(custs))
        with c4: st.metric("Top 20 Articles", f"{top20_art_pct:.0f}%")
        
        # Create full cleaned Excel matching Victor's format
        output = BytesIO()
        wb_out = openpyxl.Workbook()
        
        # 1. Summary sheet
        ws_sum = wb_out.active
        ws_sum.title = "Summary"
        ws_sum['A1'] = "HYAB MONTHLY SALES"
        ws_sum['A1'].font = Font(bold=True, size=14)
        ws_sum['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_sum['A4'] = "Total Sales (excl. VAT)"
        ws_sum['B4'] = total_art
        ws_sum['B4'].number_format = '#,##0'
        ws_sum['A5'] = "Number of Articles"
        ws_sum['B5'] = len(arts)
        ws_sum['A6'] = "Number of Customers"
        ws_sum['B6'] = len(custs)
        ws_sum['A8'] = "Top 20 Articles"
        ws_sum['B8'] = f"{top20_art_pct:.0f}% of total"
        ws_sum['A9'] = "Top 20 Customers"
        ws_sum['B9'] = f"{top20_cust_pct:.0f}% of total"
        ws_sum.column_dimensions['A'].width = 25
        ws_sum.column_dimensions['B'].width = 20
        
        # 2. Top 20 Articles sheet
        ws_t20a = wb_out.create_sheet("Top 20 Articles")
        ws_t20a['A1'] = f"TOP 20 ARTICLES ({top20_art_pct:.0f}% of total)"
        ws_t20a['A1'].font = Font(bold=True)
        ws_t20a.append([])
        ws_t20a.append(['#', 'Article No', 'Article Name', 'Amount SEK', '% of Total'])
        for i, a in enumerate(arts[:20], 1):
            ws_t20a.append([i, a['artikelnr'], a['artikelnamn'], a['summa'], f"{a['summa']/total_art*100:.1f}%" if total_art > 0 else "0%"])
        
        # 3. Top 20 Customers sheet
        ws_t20c = wb_out.create_sheet("Top 20 Customers")
        ws_t20c['A1'] = f"TOP 20 CUSTOMERS ({top20_cust_pct:.0f}% of total)"
        ws_t20c['A1'].font = Font(bold=True)
        ws_t20c.append([])
        ws_t20c.append(['#', 'Customer No', 'Customer Name', 'Amount SEK', '% of Total'])
        for i, c in enumerate(custs[:20], 1):
            ws_t20c.append([i, c['kundnr'], c['kund'], c['summa'], f"{c['summa']/total_cust*100:.1f}%" if total_cust > 0 else "0%"])
        
        # 4. Full Articles sheet
        ws_art = wb_out.create_sheet("Articles")
        ws_art.append(['Article No', 'Article Name', 'Amount SEK'])
        for a in arts:
            ws_art.append([a['artikelnr'], a['artikelnamn'], a['summa']])
        
        # 5. Full Customers sheet
        ws_cust = wb_out.create_sheet("Customers")
        ws_cust.append(['Customer No', 'Customer Name', 'Amount SEK'])
        for c in custs:
            ws_cust.append([c['kundnr'], c['kund'], c['summa']])
        
        wb_out.save(output)
        st.download_button(
            label="📥 Download Cleaned Sales Data",
            data=output.getvalue(),
            file_name=f"HYAB_Sales_{datetime.now().strftime('%Y%m')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.markdown('<p style="color:#6B7280;font-size:12px;margin-top:8px;">Contains: Summary, Top 20 Articles, Top 20 Customers, full Articles list, full Customers list</p>', unsafe_allow_html=True)

else:
    _page_title("Sales Intelligence")
    st.markdown('<p style="color:#6B7280;margin:-16px 0 20px 0;">Generate analytics dashboard from your master Sales_work_file.xlsx.</p>', unsafe_allow_html=True)
    
    st.markdown('''<div style="background:#E8F4FD;border-left:3px solid #1F4E79;padding:12px 16px;margin-bottom:16px;">
<strong>How it works:</strong>
<ul style="font-size:13px;color:#666;margin:8px 0 0 0;padding-left:20px;">
<li><strong>LTM Comparison:</strong> Compares latest LTM column vs same period last year (e.g. "LTM 25-nov" vs "LTM 24-nov")</li>
<li><strong>Churned:</strong> Customers in "bortfall" column OR had revenue last year but zero this year</li>
<li><strong>New:</strong> Customers with revenue this year but zero last year in the master file</li>
<li><strong>Growing/Declining:</strong> Year-over-year change in LTM revenue</li>
</ul>
</div>''', unsafe_allow_html=True)
    
    # API Key input for AI commentary
    _section_label("AI Commentary (optional)")
    api_key = st.text_input("Claude API Key", type="password", help="Enter your Anthropic API key to add AI-generated insights to each chart. Leave blank to skip.")
    
    _section_label("Master File")
    f = st.file_uploader("Sales_work_file.xlsx", type=['xlsx'], key="intel")
    if st.button("Generate Dashboard", type="primary", disabled=not f, use_container_width=True):
        if f:
            try:
                with st.spinner("Loading master file..."):
                    wb = openpyxl.load_workbook(f, data_only=True)
                    data = parse_master(wb)
                ltms = sorted(data['ltm_trend'].keys(), key=ltm_sort_key)
                curr = ltms[-1] if ltms else None
                prev = ltms[-13] if len(ltms) > 12 else ltms[0] if ltms else None
                if curr and prev:
                    # Prepare data for AI commentary
                    cohorts = analyze_cohorts(data, curr, prev)
                    trajectories = analyze_ltm_trajectories(data, num_periods=6)
                    
                    # Analyze industries
                    industry_data = []
                    industry_stats = defaultdict(lambda: {'curr': 0, 'prev': 0, 'churned_rev': 0, 'new_rev': 0, 'count': 0})
                    for c in data['customers']:
                        curr_v = c['ltm'].get(curr, 0)
                        prev_v = c['ltm'].get(prev, 0)
                        if curr_v == 0 and prev_v == 0:
                            continue
                        industry = classify_customer_industry(c['kund'])
                        industry_stats[industry]['curr'] += curr_v
                        industry_stats[industry]['prev'] += prev_v
                        industry_stats[industry]['count'] += 1
                        if curr_v == 0 and prev_v > 0:
                            industry_stats[industry]['churned_rev'] += prev_v
                        elif curr_v > 0 and prev_v == 0:
                            industry_stats[industry]['new_rev'] += curr_v
                    
                    for ind, stats in sorted(industry_stats.items(), key=lambda x: x[1]['curr'], reverse=True):
                        if stats['curr'] > 0 or stats['prev'] > 0:
                            change_pct = ((stats['curr'] - stats['prev']) / stats['prev'] * 100) if stats['prev'] > 0 else 0
                            industry_data.append({
                                'industry': ind,
                                'curr_ltm': stats['curr'],
                                'change_pct': change_pct,
                                'churned_rev': stats['churned_rev'],
                                'count': stats['count']
                            })
                    
                    industry_analysis = format_industry_analysis(industry_data)
                    
                    # Build chart data summary for AI
                    total_ltm = data['ltm_trend'].get(curr, 0)
                    prev_ltm = data['ltm_trend'].get(prev, 0)
                    
                    # LTM trend summary (last 6 periods)
                    recent_ltms = ltms[-6:]
                    ltm_trend_summary = "\n".join([f"- {l}: {data['ltm_trend'].get(l, 0):,.0f} SEK" for l in recent_ltms])
                    
                    # Top churned
                    top_churned = "\n".join([f"- {c['kund']}: {c['previous']:,.0f} SEK (last order: {c.get('last_month', 'unknown')})" for c in cohorts['churned'][:3]])
                    
                    # At-risk summary  
                    at_risk_summary = "\n".join([f"- {c['kund']}: down {c['decline_pct']:.0f}% over {c['consecutive_months']} months" for c in trajectories['at_risk'][:3]]) if trajectories['at_risk'] else "No customers currently at risk"
                    
                    # Top growing
                    top_growing = "\n".join([f"- {c['kund']}: +{c['change']:,.0f} SEK ({c['pct']:+.1f}%)" for c in cohorts['growing'][:3]])
                    
                    # Churn timeline summary
                    churn_tl = cohorts.get('churn_timeline', {})
                    churn_tl_sorted = sorted(churn_tl.items(), key=lambda x: x[1]['total'], reverse=True)[:3]
                    churn_timeline_summary = "\n".join([f"- {m}: {d['total']:,.0f} SEK lost ({d['count']} customers)" for m, d in churn_tl_sorted])
                    
                    chart_data = {
                        'total_ltm': total_ltm,
                        'prev_ltm': prev_ltm,
                        'yoy_chg': total_ltm - prev_ltm,
                        'yoy_pct': ((total_ltm - prev_ltm) / prev_ltm * 100) if prev_ltm > 0 else 0,
                        'active_customers': len([c for c in data['customers'] if c['ltm'].get(curr, 0) > 0]),
                        'churned_count': len(cohorts['churned']),
                        'churn_loss': sum(c['previous'] for c in cohorts['churned']),
                        'new_count': len(cohorts['new']),
                        'new_gain': sum(c['current'] for c in cohorts['new']),
                        'growing_count': len(cohorts['growing']),
                        'growth_gain': sum(c['change'] for c in cohorts['growing']),
                        'declining_count': len(cohorts['declining']),
                        'decline_loss': abs(sum(c['change'] for c in cohorts['declining'])),
                        'ltm_trend_summary': ltm_trend_summary,
                        'top_churned': top_churned,
                        'at_risk_summary': at_risk_summary,
                        'top_growing': top_growing,
                        'churn_timeline_summary': churn_timeline_summary,
                        'concentration_pct': (sum(c['ltm'].get(curr, 0) for c in sorted(data['customers'], key=lambda x: x['ltm'].get(curr, 0), reverse=True)[:20]) / total_ltm * 100) if total_ltm > 0 else 0,
                        'industry_analysis': industry_analysis,
                        'industry_data': industry_data
                    }
                    
                    # Store industry data for HTML rendering
                    data['industry_analysis'] = industry_data
                    
                    # Generate AI commentary if API key provided
                    commentary = {}
                    if api_key:
                        with st.spinner("Generating AI commentary..."):
                            commentary = generate_ai_commentary(api_key, chart_data)
                            if commentary:
                                st.success("✓ AI commentary generated")
                    
                    with st.spinner("Generating HTML dashboard..."):
                        html = generate_html(data, curr, prev, commentary)
                    st.session_state['intel_html'] = html
                    st.session_state['intel_data'] = data
                    st.session_state['intel_ltm'] = (curr, prev)
                    st.rerun()
                else: st.error("Could not find LTM data in file")
            except Exception as e:
                st.error(str(e))
                import traceback
                st.code(traceback.format_exc())
    
    if 'intel_html' in st.session_state:
        data = st.session_state['intel_data']
        curr, prev = st.session_state['intel_ltm']
        _success_banner()
        total = data['ltm_trend'].get(curr, 0)
        prev_v = data['ltm_trend'].get(prev, 0)
        yoy = ((total - prev_v) / prev_v * 100) if prev_v > 0 else 0
        c1, c2, c3, c4 = st.columns(4)
        with c1: st.metric("LTM Sales", f"{total/1e6:.1f}M SEK")
        with c2: st.metric("YoY Change", f"{yoy:+.1f}%")
        with c3: st.metric("Articles", f"{len(data['articles']):,}")
        with c4: st.metric("Customers", f"{len([c for c in data['customers'] if c['ltm'].get(curr,0)>0]):,}")
        st.markdown("---")
        st.download_button(label="📊 Download Dashboard (HTML)", data=st.session_state['intel_html'], file_name=f"HYAB_Dashboard_{datetime.now().strftime('%Y%m%d')}.html", mime="text/html", use_container_width=True)
        st.markdown('<div style="background:#F0FDF4;border-left:3px solid #16A34A;padding:12px 16px;margin-top:16px;"><strong>✓ Dashboard includes:</strong><br><span style="font-size:13px;color:#666;">• Rolling LTM trend chart<br>• Monthly sales bar chart<br>• Year-over-Year comparison by month<br>• Revenue bridge visualization<br>• Customer cohorts (Churned, Declining, Growing, New)<br>• Top 20 Customers with YoY change<br>• Top 20 Articles<br>• AI-generated insights (if API key provided)</span></div>', unsafe_allow_html=True)

st.markdown("---")

with st.expander("What does each mode do?"):
    st.markdown("""
**Order Book** — Clean and analyze your weekly order book export from Briox.
- Converts all currencies to SEK using your exchange rates
- Highlights partially invoiced orders
- Shows aging alerts for orders >90 days old
- Week-over-week comparison (if you upload last week's file)

**Sales** — Process monthly sales data for the master file.
- Parses Article and Company sheets
- Calculates Top 20 concentration
- Prepares data for pasting into Sales_work_file.xlsx

**Intelligence** — Generate a full analytics dashboard from your master file.
- Upload Sales_work_file.xlsx to create a downloadable HTML report
- Includes LTM trends, YoY comparisons, customer cohorts, Top 20 analysis
- Share the HTML with Colin or anyone — opens in any browser
    """)

st.markdown('<div style="text-align:center;color:#6B7280;font-size:12px;">HYAB Data Cleaner v3.0 · Built for Victor</div>', unsafe_allow_html=True)
